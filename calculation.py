import pandas as pd
import openpyxl

def generate_sch_from_list(lst):
  temp=[]
  for i in range(len(lst)):
    for j in range(i+1,len(lst)):
      temp.append([lst[i],lst[j],'x'])
  #print(temp)
  return temp

def Schedule_writer(schedule_dataframe,filename):
   #writer for schedule

 #delete the sheet if already exist

 workbook=openpyxl.load_workbook(filename)
 if 'Schedule' in workbook.sheetnames:
     del workbook['Schedule']
     workbook.save(filename)
 #writer.to_excel(filename,sheet_name='Schedule')
 with pd.ExcelWriter(filename,engine='openpyxl',mode='a') as wr:
                     schedule_dataframe.to_excel(wr,sheet_name='Schedule',index = False)
 return True


def PointTable_writer(PT_dataframe,filename):


 workbook=openpyxl.load_workbook(filename)
 if 'PointTable' in workbook.sheetnames:
     del workbook['PointTable']
     workbook.save(filename)
 with pd.ExcelWriter(filename,engine='openpyxl',mode='a') as wr:
                     PT_dataframe.to_excel(wr,sheet_name='PointTable',index = False)

def Leaderboard_writer(LB_dataframe,filename):


 workbook=openpyxl.load_workbook(filename)
 if 'Leaderboard' in workbook.sheetnames:
     del workbook['Leaderboard']
     workbook.save(filename)
 with pd.ExcelWriter(filename,engine='openpyxl',mode='a') as wr:
                     LB_dataframe.to_excel(wr,sheet_name='Leaderboard',index = False)

def create_new_season(filename):
 df=pd.read_excel(filename, engine ='openpyxl',sheet_name ='Groups',keep_default_na=False)
 grpdata=pd.DataFrame()
 num_of_groups= len(df.columns)
 schedule_data=[]

 player_list=[]
 print(num_of_groups)


 for i in range(num_of_groups):
    grpdata[df.columns[i]]=df.iloc[:,i]
    player_list.extend(df.iloc[:,i])
 for i in range(num_of_groups):
    schedule_data.extend(generate_sch_from_list(df.iloc[:,i]))

 writer=pd.DataFrame(schedule_data,columns=['Player1','Player2','Score'])
 Schedule_writer(writer,filename)

  #writer for pointtable
 pointtable_data=[]
 for i in range(num_of_groups):
    for j in range(len(df.index)):
      pointtable_data.extend([[df.iloc[j,i],'0','0','0','0','0',i+1]])
 writer=pd.DataFrame(pointtable_data,columns=['Player','Matches','Won','Loss','Bonus','Points','Group'])
 PointTable_writer(writer,filename)



 return num_of_groups

def update_points(filename,p1,p1points,p2,p2points,winner,bonusplayer):
  df=pd.read_excel(filename, engine ='openpyxl',sheet_name ='PointTable',keep_default_na=False)
  p1index=0
  p2index=0
  for index,row in df.iterrows():
    if(row['Player']==str(p1)):
      p1index=index
    if(row['Player']==str(p2)):
      p2index=index

  #points
  df.at[p1index,'Points']=df.at[p1index,'Points']+p1points
  df.at[p2index,'Points']=df.at[p2index,'Points']+p2points
  #matches
  df.at[p1index,'Matches']=df.at[p1index,'Matches']+1
  df.at[p2index,'Matches']=df.at[p2index,'Matches']+1
  #win and loss
  if(winner=='p1'):
    df.at[p1index,'Won']=df.at[p1index,'Won']+1
    df.at[p2index,'Loss']=df.at[p2index,'Loss']+1
  elif(winner=='p2'):
    df.at[p1index,'Loss']=df.at[p1index,'Loss']+1
    df.at[p2index,'Won']=df.at[p2index,'Won']+1
  #bonus
  if(bonusplayer=='p1'):
    df.at[p1index,'Bonus']=df.at[p1index,'Bonus']+10
  elif(bonusplayer=='p2'):
    df.at[p2index,'Bonus']=df.at[p2index,'Bonus']+10

  PointTable_writer(df,filename)

def update_points_doubles(filename,p1,p1points,p2,p2points,winner,bonusplayer,gamest1,gamest2):
  df=pd.read_excel(filename, engine ='openpyxl',sheet_name ='PointTable',keep_default_na=False)
  p1index=0
  p2index=0
  for index,row in df.iterrows():
    if(row['Team']==str(p1)):
      p1index=index
    if(row['Team']==str(p2)):
      p2index=index

  #points
  df.at[p1index,'Points']=df.at[p1index,'Points']+p1points
  df.at[p2index,'Points']=df.at[p2index,'Points']+p2points
  #matches
  df.at[p1index,'Matches']=df.at[p1index,'Matches']+1
  df.at[p2index,'Matches']=df.at[p2index,'Matches']+1
  #win and loss
  if(winner=='p1'):
    df.at[p1index,'Won']=df.at[p1index,'Won']+1
    df.at[p2index,'Loss']=df.at[p2index,'Loss']+1
  elif(winner=='p2'):
    df.at[p1index,'Loss']=df.at[p1index,'Loss']+1
    df.at[p2index,'Won']=df.at[p2index,'Won']+1
  #bonus
  if(bonusplayer=='p1'):
    df.at[p1index,'Bonus']=df.at[p1index,'Bonus']+10
  elif(bonusplayer=='p2'):
    df.at[p2index,'Bonus']=df.at[p2index,'Bonus']+10
  #percentage games
  df.at[p1index,'GamesWon']=df.at[p1index,'GamesWon']+gamest1
  df.at[p1index,'GamesTotal']=df.at[p1index,'GamesTotal']+gamest1
  df.at[p2index,'GamesWon']=df.at[p2index,'GamesWon']+gamest2
  df.at[p2index,'GamesTotal']=df.at[p2index,'GamesTotal']+gamest2

  df.at[p1index,'%games']=(df.at[p1index,'GamesWon']/df.at[p1index,'GamesTotal'])*100
  df.at[p2index,'%games']=(df.at[p2index,'GamesWon']/df.at[p2index,'GamesTotal'])*100

  PointTable_writer(df,filename)

def calc_points(p1s1,p1s2,p1s3,p2s1,p2s2,p2s3):
  p1points=0
  p2points=0
  bonus=0
  winner=''
  bonusplayer=''
  if(p1s3 or p2s3):
    bonus=10

  if((p1s1>p2s1 and p1s2>p2s2) or (p1s1>p2s1 and p1s3>p2s3) or (p1s2>p2s2 and p1s3>p2s3)):
    p1points=40
    p2points=10+bonus
    winner='p1'
    if(bonus):
      bonusplayer='p2'
  else:
    p1points=10+bonus
    p2points=40
    winner='p2'
    if(bonus):
      bonusplayer='p1'



  return ([p1points,p2points,winner,bonusplayer])


def calc_xrating(p1s1,p1s2,p1s3,p2s1,p2s2,p2s3,current_p1_xrating,current_p2_xrating):

    p1sum=p1s1+p1s2+p1s3
    p2sum=p2s1+p2s2+p2s3

    prob_p1_before=round(current_p1_xrating/(current_p1_xrating+current_p2_xrating),3)
    prob_p2_before=round(current_p2_xrating/(current_p1_xrating+current_p2_xrating),3)

    prob_p1_after=round(p1sum/(p1sum+p2sum),3)
    prob_p2_after=round(p2sum/(p1sum+p2sum),3)

    Xchange_p1=(prob_p1_after-prob_p1_before)*1000
    Xchange_p2=(prob_p2_after-prob_p2_before)*1000


    return([Xchange_p1,Xchange_p2])

  # ******* Doubles *********

def ScoreCheck(p1s1,p1s2,p1s3,p2s1,p2s2,p2s3,p1forefeit,p2forefeit):
    if((int(p1s1)<6 and int(p2s1)<6) or (int(p1s2)<6 and int(p2s2)<6)):
    #print(p1s1,p2s1)
      return False
    elif((int(p1s1)>int(p2s1)) and (int(p1s2)>int(p2s2)) and (int(p1s3)>0 or int(p2s3)>0)):
      return False
    elif((int(p1s1)<int(p2s1)) and (int(p1s2)<int(p2s2)) and (int(p1s3)>0 or int(p2s3)>0)):
      return False
    elif((int(p1s1)<int(p2s1)) and (int(p1s2)>int(p2s2)) and (int(p1s3)<6 and int(p2s3)<6)):
      return False
    elif((int(p1s1)>int(p2s1)) and (int(p1s2)<int(p2s2)) and (int(p1s3)<6 and int(p2s3)<6)):
      return False
    elif(p1forefeit and p2forefeit):
      return False
    return True
  # ******* End Doubles *********

def update_score(filename,leaderboard_filename,row_id,player1,player2,p1s1,p1s2,p1s3,p2s1,p2s2,p2s3,p1forefeit,p2forefeit):
  #update leaderboard rating and past matches
  # update Score in Schedule and pointable

  score= str(p1s1)+'-'+str(p2s1)+','+str(p1s2)+'-'+str(p2s2)+','+str(p1s3)+'-'+str(p2s3)
  df_sch=pd.read_excel(filename, engine ='openpyxl',sheet_name ='Schedule',keep_default_na=False)

  #update pointable
  if(p1forefeit or p2forefeit):
   if(p1forefeit):
     update_points(filename,player1,0,player2,40,'p2','')
     score='Forefeit by '+ str(player1)
   elif(p2forefeit):
     update_points(filename,player1,40,player2,0,'p1','')
     score='Forefeit by '+ str(player2)
   for index,row in df_sch.iterrows():
     if((row['Player1']==str(player1)) and  (row['Player2']==str(player2))):
       df_sch.at[index,'Score']= score
       Schedule_writer(df_sch,filename)
   return True

  if((int(p1s1)<6 and int(p2s1)<6) or (int(p1s2)<6 and int(p2s2)<6)):
    #print(p1s1,p2s1)
    return False
  elif((int(p1s1)>int(p2s1)) and (int(p1s2)>int(p2s2)) and (int(p1s3)>0 or int(p2s3)>0)):
    return False
  elif((int(p1s1)<int(p2s1)) and (int(p1s2)<int(p2s2)) and (int(p1s3)>0 or int(p2s3)>0)):
    return False
  elif((int(p1s1)<int(p2s1)) and (int(p1s2)>int(p2s2)) and (int(p1s3)<6 and int(p2s3)<6)):
    return False
  elif((int(p1s1)>int(p2s1)) and (int(p1s2)<int(p2s2)) and (int(p1s3)<6 and int(p2s3)<6)):
    return False

  get_points=calc_points(int(p1s1),int(p1s2),int(p1s3),int(p2s1),int(p2s2),int(p2s3))
  update_points(filename,player1,get_points[0],player2,get_points[1],str(get_points[2]),str(get_points[3]))
  #update Schedule
  for index,row in df_sch.iterrows():
    if((row['Player1']==str(player1)) and  (row['Player2']==str(player2))):
     df_sch.at[index,'Score']= score
     Schedule_writer(df_sch,filename)

  #update leaderboard
  df_LB=pd.read_excel(leaderboard_filename, engine ='openpyxl',sheet_name ='Leaderboard',keep_default_na=False)
  p1index=0
  p2index=0
  for index,row in df_LB.iterrows():
    if(row['Player']==str(player1)):
      p1index=index
    if(row['Player']==str(player2)):
      p2index=index
  new_rating=calc_xrating(int(p1s1),int(p1s2),int(p1s3),int(p2s1),int(p2s2),int(p2s3),df_LB.at[p1index,'Current Rating'],df_LB.at[p2index,'Current Rating'])

  df_LB.at[p1index,'Prev Rating']=df_LB.at[p1index,'Current Rating']
  df_LB.at[p2index,'Prev Rating']=df_LB.at[p2index,'Current Rating']
  df_LB.at[p1index,'Current Rating']=df_LB.at[p1index,'Current Rating']+new_rating[0]
  df_LB.at[p2index,'Current Rating']=df_LB.at[p2index,'Current Rating']+new_rating[1]
  Leaderboard_writer(df_LB,leaderboard_filename)
  return get_points[2]
